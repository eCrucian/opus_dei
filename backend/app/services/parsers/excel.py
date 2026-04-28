"""Excel parser — reads .xlsx/.xlsm/.xlsb and extracts sheet data, named ranges, and formulas."""
from pathlib import Path
from typing import Any, Dict, List


def parse_excel(file_path: Path) -> List[Dict[str, Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path), data_only=False, read_only=False)
    sheets = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows_data: List[List[Any]] = []
        formulas: List[str] = []

        for row in ws.iter_rows():
            row_vals = []
            for cell in row:
                val = cell.value
                row_vals.append(val)
                if isinstance(val, str) and val.startswith("="):
                    formulas.append(f"{cell.coordinate}: {val}")
            rows_data.append(row_vals)

        # Strip empty trailing rows/cols
        non_empty = [r for r in rows_data if any(v is not None for v in r)]

        sheets.append({
            "sheet_name": name,
            "rows": non_empty[:500],  # cap at 500 rows for LLM context
            "formula_count": len(formulas),
            "formulas_sample": formulas[:100],
            "dimensions": f"{ws.max_row}x{ws.max_column}",
        })

    # Named ranges
    named_ranges = {}
    for rng_name, rng_obj in wb.defined_names.items():
        try:
            named_ranges[rng_name] = str(rng_obj.attr_text)
        except Exception:
            pass

    wb.close()
    return sheets, named_ranges


def excel_to_text(file_path: Path) -> str:
    sheets, named_ranges = parse_excel(file_path)
    parts = [f"# Excel: {file_path.name}\n"]

    if named_ranges:
        parts.append("## Named Ranges")
        for k, v in named_ranges.items():
            parts.append(f"  {k}: {v}")

    for s in sheets:
        parts.append(f"\n## Sheet: {s['sheet_name']} ({s['dimensions']})")
        parts.append(f"  Fórmulas encontradas: {s['formula_count']}")
        if s["formulas_sample"]:
            parts.append("  Amostra de fórmulas:")
            for f in s["formulas_sample"][:20]:
                parts.append(f"    {f}")
        if s["rows"]:
            parts.append("  Primeiras 10 linhas:")
            for row in s["rows"][:10]:
                parts.append("    " + " | ".join(str(v) for v in row if v is not None))

    return "\n".join(parts)
